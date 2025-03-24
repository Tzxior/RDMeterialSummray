#input("中断点")

import sys
import os
import pandas as pd
import random
import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import QApplication, QWidget, QInputDialog, QVBoxLayout, QPushButton, QComboBox, QListView, QListWidget, QFileDialog, QLabel, QMessageBox, QLineEdit, QDialog, QHBoxLayout

class CustomListView(QListView):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("QListView::item { padding: 3px; }")

class CustomComboBox(QComboBox):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setView(CustomListView(self))

class DateDialog(QDialog):
    def __init__(self, initial_date, parent=None):
        super().__init__(parent)
        self.setWindowTitle('选择年月')
        self.layout = QVBoxLayout()
        self.label = QLabel(initial_date)
        self.layout.addWidget(self.label)

        self.buttonLayout = QHBoxLayout()
        self.prevButton = QPushButton('←')
        self.prevButton.clicked.connect(self.prevMonth)
        self.buttonLayout.addWidget(self.prevButton)

        self.nextButton = QPushButton('→')
        self.nextButton.clicked.connect(self.nextMonth)
        self.buttonLayout.addWidget(self.nextButton)

        self.confirmButton = QPushButton('确定')
        self.confirmButton.clicked.connect(self.accept)
        self.buttonLayout.addWidget(self.confirmButton)

        self.layout.addLayout(self.buttonLayout)
        self.setLayout(self.layout)

        self.date = pd.Timestamp(initial_date)

    def prevMonth(self):
        self.date = self.date - pd.DateOffset(months=1)
        self.label.setText(self.date.strftime('%Y-%m'))

    def nextMonth(self):
        self.date = self.date + pd.DateOffset(months=1)
        self.label.setText(self.date.strftime('%Y-%m'))

    def getDate(self):
        return self.date.strftime('%Y-%m')

class RDMaterialSummaryApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.first_import_click = True
        self.excel_date = pd.Timestamp.now().strftime('%Y-%m')
        self.prev_rd_index = 0  # 用户选择的上一个RD编号，用于回溯
        self.prev_sample_index = 0  # 用户选择的上一个样品编号，用于回溯
        self.material_array = []  # 物料数组，格式为[物料信息，可用次数，由此物料制作的样品编号，当前显示位置]
        self.sample_array = []  # 样品数组，格式为[样品编号，样品名称，样品数量，用途与处置方式，使用地点，包含物料，领用日期，处置日期]
        
    def initUI(self):
        self.setWindowTitle('研发领料汇总表')
        self.setGeometry(100, 100, 480, 480)

        self.label = QLabel('请选择本月的研发领料汇总表框架：', self)
        self.label.setGeometry(5, 0, 200, 30)
        
        self.button = QPushButton('选择excel文件', self)
        self.button.setGeometry(205, 4, 270, 22)
        self.button.clicked.connect(self.openAndCopyFile)

        self.label = QLabel('请选择要编辑的研发项目编号：', self)
        self.label.setGeometry(5, 25, 200, 30)
        
        self.comboBox1 = CustomComboBox(self)  # RD下拉框
        self.comboBox1.setGeometry(206, 30, 268, 20)
        self.comboBox1.currentIndexChanged.connect(self.handleRDChange)

        self.label = QLabel('样品编号：', self)
        self.label.setGeometry(5, 50, 80, 30)
        
        self.sampleNumberComboBox = CustomComboBox(self)  # 样品编号下拉框
        self.sampleNumberComboBox.setGeometry(65, 55, 409, 20)
        self.sampleNumberComboBox.currentIndexChanged.connect(self.handleSampleNumberChange)

        self.label = QLabel('样品名称：', self)
        self.label.setGeometry(5, 75, 80, 30)
        
        self.yhpbmkigLineEdit = QLineEdit(self)  # 样品名称文本框
        self.yhpbmkigLineEdit.setGeometry(65, 80, 284, 20)
        self.yhpbmkigLineEdit.textChanged.connect(self.checkFields)

        self.label = QLabel('数量：', self)
        self.label.setGeometry(360, 75, 60, 30)
        
        self.yhpbuullLineEdit = QLineEdit(self)  # 样品数量文本框
        self.yhpbuullLineEdit.setGeometry(400, 80, 74, 20)
        self.yhpbuullLineEdit.textChanged.connect(self.checkFields)

        self.label = QLabel('样品用途与处置方式：', self)
        self.label.setGeometry(5, 100, 120, 30)
        
        self.processComboBox = CustomComboBox(self)  # 处置方式下拉框
        self.processComboBox.setGeometry(205, 105, 269, 20)
        self.processComboBox.addItems(["样品验证-报废", "小批量验证-现场自行处理"])
        self.processComboBox.currentIndexChanged.connect(self.updateLocationComboBox)

        self.label = QLabel('样品使用地点：', self)
        self.label.setGeometry(5, 125, 120, 30)

        self.locationComboBox = CustomComboBox(self)  # 样品使用地点下拉框
        self.locationComboBox.setGeometry(205, 130, 269, 20)

        self.updateLocationComboBox()  # 初始化样品使用地点下拉框选择项

        self.addButton = QPushButton('向列表添加', self)
        self.addButton.setGeometry(5, 154, 110, 22)
        self.addButton.setEnabled(False)
        self.addButton.clicked.connect(self.addToListWidget)

        self.delButton = QPushButton('从列表删除', self)
        self.delButton.setGeometry(125, 154, 110, 22)
        self.delButton.setEnabled(False)
        self.delButton.clicked.connect(self.MinusToListWidget)

        self.plusButton = QPushButton('物料使用次数+1', self)
        self.plusButton.setGeometry(245, 154, 110, 22)
        self.plusButton.setEnabled(False)
        self.plusButton.clicked.connect(self.MaterialPlusOne)

        self.minusButton = QPushButton('物料使用次数-1', self)
        self.minusButton.setGeometry(365, 154, 110, 22)
        self.minusButton.setEnabled(False)
        self.minusButton.clicked.connect(self.MaterialMinusOne)

        self.label = QLabel('物料信息：', self)
        self.label.setGeometry(5, 175, 80, 30)

        self.comboBox2 = CustomComboBox(self)  # 物料下拉框
        self.comboBox2.setGeometry(75, 180, 399, 20)
        
        self.listWidget = QListWidget(self)  # 列表框
        self.listWidget.setGeometry(6, 205, 443, 240)
        self.listWidget.itemSelectionChanged.connect(self.updateButtonStates)

        self.upButton = QPushButton('↑', self)
        self.upButton.setGeometry(453, 204, 22, 120)
        self.upButton.setEnabled(False)
        self.upButton.clicked.connect(self.MaterialUp)

        self.downButton = QPushButton('↓', self)
        self.downButton.setGeometry(453, 326, 22, 120)
        self.downButton.setEnabled(False)
        self.downButton.clicked.connect(self.MaterialDown)

        self.saveButton = QPushButton('记录', self)
        self.saveButton.setGeometry(5, 452, 233, 22)
        self.saveButton.setEnabled(False)
        self.saveButton.clicked.connect(self.saveData)
        
        self.importButton = QPushButton('导入', self)
        self.importButton.setGeometry(242, 452, 230, 22)
        self.importButton.setEnabled(False)
        self.importButton.clicked.connect(self.importData)
        
    def openAndCopyFile(self):
        options = QFileDialog.Options()
        filePath, _ = QFileDialog.getOpenFileName(self, "打开Excel文件", "", "Excel文件 (*.xlsx);;所有文件 (*)", options=options)
        if filePath:
            date_dialog = DateDialog(pd.Timestamp.now().strftime('%Y-%m'), self)
            if date_dialog.exec_() == QDialog.Accepted:
                self.excel_date = date_dialog.getDate()
                print(f"当前处理的时间是：{self.excel_date}")
            self.filePath = filePath
            self.copyFileToDesktop()
            
    def loadExcelData(self, filePath):
        df = pd.read_excel(filePath, sheet_name=0)
        column_name = None
        for col in df.columns:
            if 'RD' in col:
                column_name = col
                break
        
        if column_name:
            unique_values = df[column_name].dropna().unique()
            rd_list = list(map(str, unique_values))
            print("排序前的RD列表:", rd_list)
            rd_list = sorted(filter(None, rd_list))
            print("排序后的RD列表:", rd_list)
            self.comboBox1.clear()
            self.comboBox1.addItems([""] + rd_list)
            self.df = df
            self.column_name = column_name
        else:
            self.comboBox1.clear()
            self.comboBox1.addItem("未找到'RD'列")
        
    def copyFileToDesktop(self):
        try:
            if hasattr(self, 'filePath'):
                desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
                self.desktop_file_path = os.path.join(desktop, f"{self.excel_date}研发领料汇总.xlsx")
                shutil.copy(self.filePath, self.desktop_file_path)

                wb = load_workbook(self.desktop_file_path, data_only=True)
                first_sheet = wb.sheetnames[0]
                sheets_to_remove = wb.sheetnames[1:]
                for sheet in sheets_to_remove:
                    wb.remove(wb[sheet])

                ws = wb[first_sheet]
                ws.title = ws['A1'].value

                # 表格第一行预处理
                merged_cells = list(ws.merged_cells)
                for merged_cell in merged_cells:
                    if merged_cell.min_row == 1:
                        ws.unmerge_cells(str(merged_cell))
                ws.delete_rows(1)

                header = {str(cell.value): col for col, cell in enumerate(ws[1], 1)}
                # 表格单价列数据预处理
                '''if '单价' in header and '金额' in header and '数量' in header:
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        if row[header['数量'] - 1].value and row[header['金额'] - 1].value:
                            row[header['单价'] - 1].value = round(row[header['金额'] - 1].value / row[header['数量'] - 1].value, 2)
                        else:
                            row[header['单价'] - 1].value = 0.00'''
                wb.save(self.desktop_file_path)

                self.material_array = []
                self.loadExcelData(self.desktop_file_path)
                for idx, row in enumerate(self.df.itertuples(index=False), 2):
                    formatted_row = self.formatRow(row, idx)
                    编号 = str(getattr(row, '料件编号', ''))
                    品名 = str(getattr(row, '品名', ''))
                    规格 = str(getattr(row, '规格', ''))
                    单位 = str(getattr(row, '发料单位', ''))
                    数量 = str(getattr(row, '发料数量', ''))
                    金额 = str(getattr(row, '金额', ''))
                    RDid = str(getattr(row, 'RD', ''))
                    self.material_array.append([formatted_row, 1, "", "其他", 编号, 品名, 规格, 单位, 数量, 金额, RDid])
                    
                self.addButton.setEnabled(True)

        except Exception as e:
            self.showError(str(e))
    
    def populateSecondComboBox(self):
        selected_value = self.comboBox1.currentText()
        self.comboBox2.clear()
        if selected_value:
            for item in self.material_array:
                #####print(f"{item[0]}{item[1]}{item[2]}{item[3]}")#####调试信息
                if item[1] - item[2].count('#') > 0 and item[3] == "其他":
                    row_num = int(item[0].split('.')[0])
                    if item[1] > 1:
                        print(f"{item[1]}{item[2]}")
                    if self.df.at[row_num - 2, self.column_name] == selected_value:
                        self.comboBox2.addItem(f"{item[0]}<{item[2].count('#')}/{item[1]}>")
    
    def clearListWidget(self):
        self.listWidget.clear()
    
    def updateSampleNumberComboBox(self):
        selected_value = self.comboBox1.currentText()
        self.sampleNumberComboBox.clear()
        if selected_value:
            # 获取当前RD项目的所有样品编号
            sample_numbers = [sample[0] for sample in self.sample_array if sample[0].startswith(selected_value)]
            sample_numbers.sort()

            # 找到下一个样品编号
            if sample_numbers:
                last_sample_number = sample_numbers[-1]
                last_suffix = int(last_sample_number.split('/')[-1])
                next_suffix = str(last_suffix + 1).zfill(3)
                next_sample_number = f"{selected_value}/{next_suffix}"
            else:
                next_sample_number = f"{selected_value}/{self.excel_date.replace('-', '')}001"

            # 添加所有样品编号和下一个样品编号到下拉框
            self.sampleNumberComboBox.addItems(sample_numbers + [next_sample_number])
        else:
            self.sampleNumberComboBox.clear()

    def updateLocationComboBox(self):
        selected_process = self.processComboBox.currentText()
        self.locationComboBox.clear()
        if selected_process == "样品验证-报废":
            self.locationComboBox.addItems(["研发中心样品测试区域", "中心实验室", "试验塔"])
            self.locationComboBox.setCurrentIndex(0)
            self.locationComboBox.setEnabled(True)
        elif selected_process == "小批量验证-现场自行处理":
            self.locationComboBox.addItem("手动输入项目名称。。。")
            self.locationComboBox.setCurrentIndex(0)
            self.locationComboBox.setEnabled(False)

    def addToListWidget(self):
        selected_value = self.comboBox2.currentText().split('<')[0]
        sample_number = self.sampleNumberComboBox.currentText()
        if selected_value and not self.isSampleNumberUsed(selected_value, sample_number):
            self.listWidget.addItem(selected_value)
            for item in self.material_array:
                if item[0] == selected_value:
                    item[3] = "列表框"
            self.populateSecondComboBox()
            self.updateButtonStates()
            self.checkFields()

    def MinusToListWidget(self):
        selected_items = self.listWidget.selectedItems()
        if selected_items:
            for item in selected_items:
                item_text = item.text()
                self.listWidget.takeItem(self.listWidget.row(item))
                for material in self.material_array:
                    if material[0] == item_text:
                        material[3] = "其他"
            self.populateSecondComboBox()
            self.updateButtonStates()
            self.checkFields()
    
    def isSampleNumberUsed(self, material, sample_number):
        for item in self.material_array:
            if item[0] == material and f"#{sample_number}" in item[2]:
                return True
        return False

    def handleRDChange(self):
        print("RD编号发生变更")
        if self.listWidget.count() > 0:
            current_materials = [self.listWidget.item(i).text() for i in range(self.listWidget.count())]
            selected_sample_number = self.sampleNumberComboBox.currentText()
            existing_sample = next((item for item in self.sample_array if item[0] == selected_sample_number), None)
            if existing_sample:
                expected_materials = [material for material in existing_sample[5]]
                if set(current_materials) != set(expected_materials):
                    response = QMessageBox.warning(self, "警告", "有数据未保存，是否切换RD编号", QMessageBox.Yes | QMessageBox.No)
                    if response == QMessageBox.No:
                        self.comboBox1.blockSignals(True)
                        self.comboBox1.setCurrentIndex(self.prev_rd_index)
                        self.comboBox1.blockSignals(False)
                        return
                    else:
                        self.clearListWidget()
                        self.clearMaterialArraySampleNumbers()
                        
            for item in current_materials:
                for material in self.material_array:
                    if material[0] == item:
                        material[3] = "其他"

        self.prev_rd_index = self.comboBox1.currentIndex()
        self.yhpbmkigLineEdit.clear()
        self.yhpbuullLineEdit.clear()
        self.processComboBox.setCurrentText("样品验证-报废")
        self.locationComboBox.setCurrentText("研发中心样品测试区域")
        self.populateSecondComboBox()
        self.updateSampleNumberComboBox()
        self.checkFields()

    def handleSampleNumberChange(self):
        print("样品编号发生变更")
        if self.listWidget.count() > 0:
            selected_sample_number = self.sampleNumberComboBox.currentText()
            existing_sample = next((item for item in self.sample_array if item[0] == selected_sample_number), None)
            current_materials = [self.listWidget.item(i).text() for i in range(self.listWidget.count())]
            if existing_sample:
                expected_materials = [material for material in existing_sample[5]]
                if set(current_materials) != set(expected_materials):
                    response = QMessageBox.warning(self, "警告", "有数据未保存，是否切换样品编号", QMessageBox.Yes | QMessageBox.No)
                    if response == QMessageBox.No:
                        self.sampleNumberComboBox.blockSignals(True)
                        self.sampleNumberComboBox.setCurrentIndex(self.prev_sample_index)
                        self.sampleNumberComboBox.blockSignals(False)
                        return
                    else:
                        for item in current_materials:
                            for material in self.material_array:
                                if material[0] == item:
                                    material[3] = "其他"
                        self.clearListWidget()
                        self.clearMaterialArraySampleNumbers()
                        
            for item in current_materials:
                for material in self.material_array:
                    if material[0] == item:
                        material[3] = "其他"                    

        self.prev_sample_index = self.sampleNumberComboBox.currentIndex()
        selected_sample_number = self.sampleNumberComboBox.currentText()
        existing_sample = next((item for item in self.sample_array if item[0] == selected_sample_number), None)
        if existing_sample:
            self.yhpbmkigLineEdit.setText(existing_sample[1])
            self.yhpbuullLineEdit.setText(existing_sample[2])
            self.processComboBox.setCurrentText(existing_sample[3])
            self.locationComboBox.setCurrentText(existing_sample[4])
            self.listWidget.clear()
            for material in existing_sample[5]:
                material_text = next((mat[0] for mat in self.material_array if mat[0].split('.')[0] == material), "")
                self.listWidget.addItem(material_text)
                for item in self.material_array:
                    if item[0] == material_text:
                        item[3] = "列表框"
        else:
            self.yhpbmkigLineEdit.clear()
            self.yhpbuullLineEdit.clear()
            self.processComboBox.setCurrentText("样品验证-报废")
            self.locationComboBox.setCurrentText("研发中心样品测试区域")
            self.clearListWidget()
        self.checkFields()
        self.populateSecondComboBox()

    def clearMaterialArraySampleNumbers(self):
        sample_number = self.sampleNumberComboBox.currentText()
        for item in self.material_array:
            item[2] = item[2].replace(f"#{sample_number}", "")

    def MaterialPlusOne(self):
        selected_items = self.listWidget.selectedItems()
        if selected_items:
            for item in selected_items:
                item_text = item.text()
                for material in self.material_array:
                    if material[0] == item_text and material[1] < 5:
                        material[1] += 1
                        print(f"Material Array after +1: {material}")
            self.populateSecondComboBox()
            self.updateButtonStates()
            self.checkFields()

    def MaterialMinusOne(self):
        selected_items = self.listWidget.selectedItems()
        if selected_items:
            for item in selected_items:
                item_text = item.text()
                for material in self.material_array:
                    if material[0] == item_text and material[1] > material[2].count('#'):
                        material[1] -= 1
                        print(f"Material Array after -1: {material}")
            self.populateSecondComboBox()
            self.updateButtonStates()
            self.checkFields()

    def MaterialUp(self):
        selected_items = self.listWidget.selectedItems()
        if selected_items:
            for item in selected_items:
                row = self.listWidget.row(item)
                if row > 0:
                    self.listWidget.takeItem(row)
                    self.listWidget.insertItem(row - 1, item)
                    self.listWidget.setCurrentItem(item)
        self.checkFields()

    def MaterialDown(self):
        selected_items = self.listWidget.selectedItems()
        if selected_items:
            for item in selected_items:
                row = self.listWidget.row(item)
                if row < self.listWidget.count() - 1:
                    self.listWidget.takeItem(row)
                    self.listWidget.insertItem(row + 1, item)
                    self.listWidget.setCurrentItem(item)
        self.checkFields()

    def formatRow(self, row, idx):
        序号 = f"{idx}"
        品名 = str(getattr(row, '品名', ''))[:30]
        规格 = str(getattr(row, '规格', ''))[:20]
        数量 = str(getattr(row, '发料数量', ''))
        单位 = str(getattr(row, '发料单位', ''))
        return f"{序号}.{品名}({规格})_{数量} {单位}"
    
    def saveData(self):
        try:
            if self.locationComboBox.currentText() == "手动输入项目名称。。。":
                project_name, ok = QInputDialog.getText(self, "输入项目名称", "请输入项目名称：")
                if ok and project_name:
                    self.locationComboBox.setItemText(0, project_name)
                else:
                    return

            sample_number = self.sampleNumberComboBox.currentText()
            sample_name = self.yhpbmkigLineEdit.text()
            sample_quantity = self.yhpbuullLineEdit.text()
            process_mode = self.processComboBox.currentText()
            usage_location = self.locationComboBox.currentText()
            material_list = [self.listWidget.item(i).text().split('.')[0] for i in range(self.listWidget.count())]

            # 3. 在两个列表中记录数据
            existing_sample = next((item for item in self.sample_array if item[0] == sample_number), None)
            if existing_sample:
                # 如果数据已存在
                response = QMessageBox.question(self, "覆盖确认", "样品编号已存在，是否覆盖原有记录？", QMessageBox.Yes | QMessageBox.No)
                if response == QMessageBox.Yes:
                    # 如果数据已存在，且确认覆盖
                    original_materials = existing_sample[5]
                    changed_materials = set(original_materials).union(set(material_list)) - set(original_materials).intersection(set(material_list))
                    
                    for material in changed_materials:
                        material_item = next((item for item in self.material_array if item[0].split('.')[0] == material), None)
                        if material_item:
                            print(f"物料变化前: {material_item}")
                            if material in original_materials and material not in material_list:
                                # 只存在于原数据，从[由此物料制作的样品编号]项中删除当前项目编号
                                material_item[2] = material_item[2].replace(f"#{sample_number}", "")
                            elif material in material_list and material not in original_materials:
                                # 只存在于新数据，在[由此物料制作的样品编号]项中新增当前项目编号
                                material_item[2] += f"#{sample_number}"
                            print(f"物料变化后: {material_item}")

                    # 更新 sample_array 中的对应样品信息
                    existing_sample[1] = sample_name
                    existing_sample[2] = sample_quantity
                    existing_sample[3] = process_mode
                    existing_sample[4] = usage_location
                    existing_sample[5] = material_list
                
                else:
                    # 如果数据已存在，且取消覆盖（这部分正确）
                    self.clearListWidget()
                    self.sampleNumberComboBox.setCurrentIndex(self.sampleNumberComboBox.count() - 1)
                    return
            else:
                # 如果数据不存在
                rd_project = sample_number.split('/')[0]
                sample_suffix = int(sample_number.split('/')[-1])
                if sample_number.endswith("001"):
                    year_month = self.excel_date
                    day = random.randint(1, 3)
                    sample_date = pd.Timestamp(f"{year_month}-{day:02d}")
                    print(f"样品{sample_number}分配到的日期是：{sample_date}")#
                else:
                    prev_sample_number = f"{rd_project}/{self.excel_date.replace('-', '')}{str(sample_suffix - 1).zfill(3)}"
                    prev_sample_date = next((item[6] for item in self.sample_array if item[0] == prev_sample_number), None)
                    if prev_sample_date:
                        sample_date = pd.Timestamp(prev_sample_date) + pd.DateOffset(days=random.randint(1, 4))
                    else:
                        sample_date = pd.Timestamp.now()
                    print(f"样品{sample_number}分配到的日期是：{sample_date}")#

                if usage_location == "中心实验室":
                    disposal_date = sample_date + pd.DateOffset(months=3)
                elif usage_location == "研发中心样品测试区域":
                    disposal_date = sample_date + pd.DateOffset(months=2)
                elif usage_location == "试验塔":
                    disposal_date = sample_date + pd.DateOffset(months=6)
                else:
                    disposal_date = None

                # 记录新的样品信息
                sample_array_element = [
                    sample_number,
                    sample_name,
                    sample_quantity,
                    process_mode,
                    usage_location,
                    material_list,
                    sample_date.strftime('%Y-%m-%d'),
                    disposal_date.strftime('%Y-%m-%d') if disposal_date else ''
                ]
                self.sample_array.append(sample_array_element)

                # 更新 material_array，将当前样品编号添加到对应物料的样品编号列表中
                for item in material_list:
                    material_item = next((material for material in self.material_array if material[0].split('.')[0] == item), None)
                    if material_item:
                        print(f"物料变化前: {material_item}")
                        material_item[2] += f"#{sample_number}"
                        print(f"物料变化后: {material_item}")

            self.yhpbmkigLineEdit.clear()
            self.yhpbuullLineEdit.clear()
            self.listWidget.clear()
            self.sampleNumberComboBox.setCurrentIndex(0)
            self.checkFields()
            self.updateSampleNumberComboBox()###########

            print("保存的样品信息:", sample_array_element)

            current_sample_index = self.sampleNumberComboBox.currentIndex()
            next_sample_index = current_sample_index + 1
            while next_sample_index < self.sampleNumberComboBox.count():
                next_sample_number = self.sampleNumberComboBox.itemText(next_sample_index)
                if not any(item[0] == next_sample_number for item in self.sample_array):
                    self.sampleNumberComboBox.setCurrentIndex(next_sample_index)
                    break
                next_sample_index += 1

        except Exception as e:
            self.showError(str(e))
    
    def importData(self):
        try:
            if not self.desktop_file_path:
                raise Exception("没有找到有效的Excel文件路径")

            # 1. 读取第一个工作表的名称
            wb = load_workbook(self.desktop_file_path)
            first_sheet_name = wb.sheetnames[0]
            sheetname_2 = f"{first_sheet_name}用途及最终去向（研发）"

            # 2. 检查是否存在同名工作表
            if sheetname_2 in wb.sheetnames:
                response = QMessageBox.question(self, "确认", f"{sheetname_2} 已存在，是否使用新数据覆盖？", QMessageBox.Yes | QMessageBox.No)
                if response == QMessageBox.No:
                    return

                sheet = wb[sheetname_2]
                # 取消所有合并单元格
                merged_cells = list(sheet.merged_cells)
                for merged_cell in merged_cells:
                    sheet.unmerge_cells(str(merged_cell))
                # 清空工作表
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.value = None
            else:
                # 新建工作表
                sheet = wb.create_sheet(sheetname_2)

            # 3. 设置工作表格式
            for row in sheet.iter_rows():
                for cell in row:
                    cell.fill = None
                    cell.font = Font(name='宋体', size=10, color='000000')
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            # 4. 设置表头
            headers = ["料件编号", "品名", "规格", "发料单位", "发料数量", "金额", "RD", "用途", "最终去向", "样品编号"]
            for col, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col)
                cell.value = header

            # 5. 输入所有的RD编号到L列，并初始化RD_row和editing_row
            rd_numbers = list(set([sample[0].split('/')[0] for sample in self.sample_array]))
            rd_numbers.sort()
            RD_row = []
            for i, rd in enumerate(rd_numbers, start=1):
                sheet.cell(row=i, column=12).value = rd
                samples = [sample for sample in self.sample_array if sample[0].startswith(rd)]
                RD_row.append([rd, samples, i])
            editing_row = 2

            # 6. 循环处理每个RD编号
            for rd, samples, start_row in RD_row:
                for sample in samples:
                    self.yhpblkysdj(sheet, sample, editing_row)
                    wb.save(self.desktop_file_path)
                    editing_row += max(5, len(sample[5]) + 1)
                    #input("按下回车键继续...")

        except Exception as e:
            self.showError(str(e))

    def yhpblkysdj(self, sheet, sample, editing_row):
        # 在S列，editing_row+1行的单元格，输入“YPLY”+self.sample_array的第六项[领用日期]（格式改为yyyymmdd）+“01”
        sheet.cell(row=editing_row + 1, column=19).value = "YPLY" + sample[6].replace("-", "") + "01"
        sheet.cell(row=editing_row + 1, column=19).alignment = Alignment(horizontal='right')

        # 把L列editing_row+1行，至R列editing_row+1行的单元格合并，单元格内容改为"编号"
        sheet.merge_cells(start_row=editing_row + 1, start_column=12, end_row=editing_row + 1, end_column=18)
        sheet.cell(row=editing_row + 1, column=12).value = "编号"
        sheet.cell(row=editing_row + 1, column=12).alignment = Alignment(horizontal='right')

        # 把L列editing_row行，至S列editing_row行的单元格合并，单元格内容改为“样品领用单”
        sheet.merge_cells(start_row=editing_row, start_column=12, end_row=editing_row, end_column=19)
        cell = sheet.cell(row=editing_row, column=12)
        cell.value = "样品领用单"
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)

        # 设置表头
        headers = ["样品编号", "样品名称", "数量", "用途", "地点", "样品管理人", "样品领用人/日期", "样品领用审批/日期"]
        for col, header in enumerate(headers, start=12):
            cell = sheet.cell(row=editing_row + 2, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 填写样品信息
        sample_info = [sample[0], sample[1], sample[2], sample[3].split('-')[0], sample[4], "沈利贞", f"谈震潇/{sample[6]}", f"茹晓英/{sample[6]}"]
        for col, info in enumerate(sample_info, start=12):
            sheet.cell(row=editing_row + 3, column=col).value = info

        # 填写物料信息
        self.yhpbsoyswulnlpbn(sheet, sample, editing_row)

        # 用途为“样品验证”的样品，需要报废处理
        if sample[3].split('-')[0] == "样品验证":
            self.yhpbiuvidj(sheet, sample, editing_row)

    def yhpbiuvidj(self, sheet, sample, editing_row):
        # 在Z列，editing_row+1行的单元格，输入“YPCZ”+self.sample_array的第六项[领用日期]（格式改为yyyymmdd）+“01”，此单元格改为右对齐
        sheet.cell(row=editing_row + 1, column=26).value = "YPCZ" + sample[6].replace("-", "") + "01"
        sheet.cell(row=editing_row + 1, column=26).alignment = Alignment(horizontal='right')

        # 把U列editing_row+1行，至Y列editing_row+1行的单元格合并，单元格内容改为"编号"，此单元格改为右对齐
        sheet.merge_cells(start_row=editing_row + 1, start_column=21, end_row=editing_row + 1, end_column=25)
        sheet.cell(row=editing_row + 1, column=21).value = "编号"
        sheet.cell(row=editing_row + 1, column=21).alignment = Alignment(horizontal='right')

        # 把U列editing_row行，至Z列editing_row行的单元格合并，单元格内容改为“样品处置单”，此单元格改为左右居中、加粗
        sheet.merge_cells(start_row=editing_row, start_column=21, end_row=editing_row, end_column=26)
        cell = sheet.cell(row=editing_row, column=21)
        cell.value = "样品处置单"
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)

        # 设置表头
        headers = ["样品编号", "样品名称", "数量", "处置方式", "样品处置申请人/日期", "样品处置审批/日期"]
        for col, header in enumerate(headers, start=21):
            cell = sheet.cell(row=editing_row + 2, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 填写样品处置信息
        disposal_info = [sample[0], sample[1], sample[2], "报废", f"谈震潇/{sample[7]}", f"茹晓英/{sample[7]}"]
        for col, info in enumerate(disposal_info, start=21):
            sheet.cell(row=editing_row + 3, column=col).value = info

    def yhpbsoyswulnlpbn(self, sheet, sample, editing_row):
        row = editing_row + 3
        for i, material in enumerate(self.material_array):
            if sample[0] in material[2]:
                sheet.cell(row=row, column=1).value = material[4]
                sheet.cell(row=row, column=2).value = material[5]
                sheet.cell(row=row, column=3).value = material[6]
                sheet.cell(row=row, column=4).value = material[7]
                sheet.cell(row=row, column=5).value = material[8]
                sheet.cell(row=row, column=6).value = material[9]
                sheet.cell(row=row, column=7).value = material[10]
                sheet.cell(row=row, column=8).value = sample[3].split('-')[0]
                sheet.cell(row=row, column=9).value = sample[3].split('-')[1]
                sheet.cell(row=row, column=10).value = sample[0]
                row += 1
    
    def copy_row(self, ws, row_data):
        row_num = ws.max_row + 1
        for idx, value in enumerate(row_data):
            col_letter = get_column_letter(idx + 1)
            cell = ws[f"{col_letter}{row_num}"]
            cell.value = value
    
    def showError(self, message):
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Critical)
        msgBox.setText("发生错误")
        msgBox.setInformativeText(message)
        msgBox.setWindowTitle("错误")
        msgBox.exec_()

    def checkFields(self):
        if (self.listWidget.count() > 0 and self.sampleNumberComboBox.currentText() and 
            self.yhpbmkigLineEdit.text() and self.yhpbuullLineEdit.text()):
            self.saveButton.setEnabled(True)
        else:
            self.saveButton.setEnabled(False)

        all_materials_satisfied = all(item[1] - item[2].count('#') == 0 for item in self.material_array)
        #self.importButton.setEnabled(all_materials_satisfied)
        self.importButton.setEnabled(True) # ，暂时不考虑全部物料分配完毕

    def updateButtonStates(self):
        has_selection = len(self.listWidget.selectedItems()) > 0
        self.delButton.setEnabled(has_selection)
        self.plusButton.setEnabled(has_selection and self.getSelectedMaterialCount() < 5)
        self.minusButton.setEnabled(has_selection and self.getSelectedMaterialCount() > 1)
        self.upButton.setEnabled(has_selection and self.listWidget.currentRow() > 0)
        self.downButton.setEnabled(has_selection and self.listWidget.currentRow() < self.listWidget.count() - 1)

    def getSelectedMaterialCount(self):
        selected_items = self.listWidget.selectedItems()
        if selected_items:
            item_text = selected_items[0].text()
            for material in self.material_array:
                if material[0] == item_text:
                    return material[1]
        return 0

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = RDMaterialSummaryApp()
    ex.show()
    sys.exit(app.exec_())
