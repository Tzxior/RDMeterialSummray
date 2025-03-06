import calendar
import openpyxl
import os
import pandas as pd
import random
import shutil
import sys
import time
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill
from PyQt5.QtCore import Qt, pyqtSignal, pyqtSlot
from PyQt5.QtGui import QIntValidator
from PyQt5.QtWidgets import QApplication, QCheckBox, QComboBox, QDialog, QGridLayout, QHBoxLayout, QInputDialog, QLabel, QLineEdit, QMainWindow, QMessageBox, QProgressBar, QPushButton, QVBoxLayout, QWidget
from tkinter import Tk, filedialog

class CalendarApp(QMainWindow):
    def __init__(self, parent=None, start_date=None, end_date=None):
        super().__init__(parent)
        self.start_date = start_date
        self.end_date = end_date
        self.initUI()

    def initUI(self):
        self.setWindowTitle('定制日历')
        self.setGeometry(100, 100, 600, 400)
        self.centralWidget = QWidget(self)
        self.setCentralWidget(self.centralWidget)
        self.layout = QVBoxLayout(self.centralWidget)

        # 显示日历
        self.display_calendar(self.start_date, self.end_date)

        # 增加传递按钮
        self.transfer_button = QPushButton("传递")
        self.transfer_button.clicked.connect(self.transfer_data)
        self.layout.addWidget(self.transfer_button)

    def display_calendar(self, start_date, end_date):
        cal = calendar.Calendar()

        # 创建网格布局
        calendar_layout = QGridLayout()
        weekdays = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]
        for i, day in enumerate(weekdays):
            calendar_layout.addWidget(QLabel(day), 0, i)

        # 填充每一天的日期和复选框
        current_date = start_date
        row_num = 1
        self.working_days = []  # 重置工作日列表
        self.checkboxes = []  # 存储复选框

        while current_date <= end_date:
            col_num = current_date.weekday()

            day_label = QLabel(current_date.strftime("%Y.%m.%d"))
            checkbox = QCheckBox()
            if col_num in (5, 6):
                checkbox.setChecked(False)
            else:
                checkbox.setChecked(True)

            layout = QVBoxLayout()
            layout.addWidget(day_label)
            layout.addWidget(checkbox)
            cell_widget = QWidget()
            cell_widget.setLayout(layout)
            calendar_layout.addWidget(cell_widget, row_num, col_num)

            # 存储复选框和日期
            self.checkboxes.append((current_date, checkbox))

            current_date += timedelta(days=1)
            if col_num == 6:  # 新的一行
                row_num += 1

        # 设置布局
        self.layout.addLayout(calendar_layout)

    def transfer_data(self):
        try:
            # 读取每个复选框的状态
            self.working_days = [(date, checkbox.isChecked()) for date, checkbox in self.checkboxes]

            # 打印调试信息
            for date, is_working in self.working_days:
                status = "工作日" if is_working else "休息日"
                print(f"日期: {date.strftime('%Y.%m.%d')} - {status}")

            parent = self.parent()
            if parent and parent.start_date_str:
                # 获取起始日期并计算结束日期
                start_date = datetime.strptime(parent.start_date_str, '%Y.%m.%d')
                if start_date.month == 12:
                    end_date = start_date.replace(year=start_date.year + 1, month=1) - timedelta(days=1)
                else:
                    end_date = start_date.replace(month=start_date.month + 1) - timedelta(days=1)

                if parent.desktop_full_path and parent.start_date_str:
                    # 创建工作日字典
                    working_days_dict = {date.strftime('%Y.%m.%d'): is_working for date, is_working in self.working_days}

                    # 在读取到需要的数据后立即关闭日历窗口
                    self.close()
                    parent.tabulate_all_work()
                    
                    # 传递数据并更新Excel文件内容
                    parent.update_excel_with_calendar_data(parent.desktop_full_path, parent.start_date_str, parent.end_date_str, working_days_dict)
                    parent.list_sheets(parent.desktop_full_path)
                    parent.populate_person_selector(parent.desktop_full_path)
                else:
                    print("路径或日期未定义")
        except Exception as e:
            print(f"transfer_data 发生错误: {e}")

class ProjectInputWindow(QWidget):
    days_updated = pyqtSignal(list)  # 定义信号

    def __init__(self, projects, days, total_days):
        super().__init__()
        self.projects = projects
        self.days = days
        self.total_days = total_days
        self.initUI()

    def initUI(self):
        self.setWindowTitle('工时分配')
        self.resize(240, 300)
        self.layout = QVBoxLayout()

        # 添加项目输入框
        for project, day in zip(self.projects, self.days):
            layout = QHBoxLayout()
            label = QLabel(str(project))
            input_field = QLineEdit()
            input_field.setText(str(day))
            input_field.setValidator(QIntValidator(0, 100))
            layout.addWidget(label)
            layout.addWidget(input_field)
            self.layout.addLayout(layout)

        # 添加确认并应用按钮
        self.apply_button = QPushButton('确认并应用')
        self.apply_button.clicked.connect(self.apply_changes)
        self.layout.addWidget(self.apply_button)

        self.setLayout(self.layout)

    def apply_changes(self):
        entered_days = [int(input_field.text()) for input_field in self.findChildren(QLineEdit)]
        total_entered_days = sum(entered_days)

        if total_entered_days > self.total_days:
            QMessageBox.critical(self, '输入错误', f'工时超过，总天数：{self.total_days} 天，多{total_entered_days - self.total_days}天')
        else:
            if total_entered_days < self.total_days:
                QMessageBox.warning(self, '输入警告', f'工时不满，总天数：{self.total_days} 天，少{self.total_days - total_entered_days}天')
            print(f'entered_days: {entered_days}')
            self.days_updated.emit(entered_days)  # 发出信号
            self.close()

class ExcelApp(QWidget):
    def __init__(self):
        super().__init__()
        self.project_list = []
        self.default_days = []
        self.working_days_dict = {}  # 初始化工作日字典
        self.all_project_data = {}  # 保存所有人的项目和分配天数
        self.person_project_days = {}  # 保存每个人负责项目的天数分配
        self.desktop_full_path = None
        self.start_date_str = None  # 保存起始日期字符串
        self.excel_info = None  # 保存Excel信息
        self.work_schedule = {}  # 初始化工作计划
        self.xl = None  # 保存加载的Excel工作簿
        self.company = None  # "森赫"或"联合"
        self.file_path = ""  # 保存Excel模板的路径

        print(f"1 {self.file_path}")
        self.file_path = ""
        
        if len(sys.argv) > 1:
            self.file_path = sys.argv[1]
            #_, file_extension = os.path.splitext(file_path)
        print(f"2 {self.file_path}")
                
        # 初始化主布局和项目输入布局
        self.mainLayout = QVBoxLayout()
        self.projectInputLayout = QVBoxLayout()
        
        self.initUI()

    def initUI(self):
        self.setWindowTitle('研发人员工时安排')
        self.resize(360, 270)

        # 设置主窗口布局
        self.setLayout(self.mainLayout)

        # 日期提示词
        self.tsLabel = QLabel('请输入开始日期：', self)
        self.tsLabel.setGeometry(20, 0, 300, 30)
        self.tsLabel = QLabel('请输入结束日期：', self)
        self.tsLabel.setGeometry(20, 15, 300, 30)
        
        # 开始日期输入框
        self.startdateInput = QLineEdit(self)
        self.startdateInput.setGeometry(120, 8, 120, 16)
        today = datetime.today()
        if today.day < 10:
            if today.month < 3:
                default_date = today.replace(year=today.year - 1, month=today.month+10, day=26)
            else:
                default_date = today.replace(month=today.month - 2, day=26)
        else:
            default_date = today.replace(month=today.month - 1, day=26)
        self.startdateInput.setText(default_date.strftime('%Y.%m.%d'))
        self.startdateInput.setPlaceholderText('YYYY.MM.DD')

        # 结束日期输入框
        self.enddateInput = QLineEdit(self)
        self.enddateInput.setGeometry(120, 23, 120, 16)
        today = datetime.today()
        if today.day < 10:
            if today.month < 2:
                default_date = today.replace(year=today.year - 1, month=today.month+11, day=25)
            else:
                default_date = today.replace(month=today.month - 1, day=25)
        else:
            default_date = today.replace(day=25)
        self.enddateInput.setText(default_date.strftime('%Y.%m.%d'))
        self.enddateInput.setPlaceholderText('YYYY.MM.DD')

        # 每个工作表对应的项目
        self.sheetSelector = QComboBox(self)
        self.sheetSelector.setGeometry(20, 40, 95, 20)
        self.sheetSelector.currentTextChanged.connect(self.show_a2_content)
        self.a2tsLabel = QLabel('A2单元格内容：', self)
        self.a2tsLabel.setGeometry(120, 40, 300, 20)
        self.a2Label = QLabel('', self)
        self.a2Label.setGeometry(20, 58, 320, 30)

        # 每个人负责/参与的项目
        self.personSelector = QComboBox(self)
        self.personSelector.setGeometry(20, 90, 65, 20)
        self.personSelector.currentTextChanged.connect(self.tabulate_every1s_work)
        self.persontstsLabel = QLabel('参与的项目：', self)
        self.persontstsLabel.setGeometry(90, 90, 300, 20)
        self.personLabel = QLabel('', self)
        self.personLabel.setGeometry(20, 115, 320, 300)
        self.personLabel.setWordWrap(True)
        self.personLabel.setAlignment(Qt.AlignLeft | Qt.AlignTop)

        # 添加天数分配调整按钮
        self.allocateButton = QPushButton('天数分配调整', self)
        self.allocateButton.setGeometry(200, 90, 100, 20)
        self.allocateButton.clicked.connect(self.open_project_input_window)
        self.allocateButton.setEnabled(False)

        # 按钮三步走
        self.copyButton = QPushButton('生成框架', self)
        self.copyButton.setGeometry(10, 200, 80, 20)
        self.copyButton.clicked.connect(self.get_start_date)
        self.copyButton.setEnabled(True)
        
        self.arrangeButton = QPushButton('工时分配', self)
        self.arrangeButton.setGeometry(10, 220, 80, 20)
        self.arrangeButton.clicked.connect(self.assign_work_schedule)
        self.arrangeButton.setEnabled(False)

        self.writeButton = QPushButton('数据写入', self)
        self.writeButton.setGeometry(10, 240, 80, 20)
        self.writeButton.clicked.connect(self.write_schedule_2_excel)
        self.writeButton.setEnabled(False)
        
        # 添加进度条
        self.progressBar = QProgressBar(self)
        self.progressBar.setGeometry(100, 203, 240, 12)
        self.progressBar.setMinimum(0)
        self.progressBar.setMaximum(100)
        self.progressBar2 = QProgressBar(self)
        self.progressBar2.setGeometry(100, 223, 240, 12)
        self.progressBar2.setMinimum(0)
        self.progressBar2.setMaximum(100)
        self.progressBar3 = QProgressBar(self)
        self.progressBar3.setGeometry(100, 243, 240, 12)
        self.progressBar3.setMinimum(0)
        self.progressBar3.setMaximum(100)

        self.show()

    def get_start_date(self):  # 读取初始日期
        print(f"3 {self.file_path}")
        self.copyButton.setEnabled(False)
        try:
            start_date_str = self.startdateInput.text().strip()
            end_date_str = self.enddateInput.text().strip()
            if self.validate_date(start_date_str, end_date_str):
                start_date = datetime.strptime(start_date_str, "%Y.%m.%d")
                end_date = datetime.strptime(end_date_str, "%Y.%m.%d")
                if not self.is_reasonable_date(start_date, end_date):
                    QMessageBox.critical(self, '输入错误', '请输入合理的日期格式：YYYY.MM.DD。')
                    return
                self.startdateInput.setDisabled(True)  # 确认输入后禁用输入框
                self.enddateInput.setDisabled(True)
                self.start_date_str = start_date_str  # 保存起始日期字符串
                self.end_date_str = end_date_str

                # 复制并重命名模板文件
                self.copy_rename_list_sheets(start_date, end_date, self.file_path)

                # 创建并显示 CalendarApp 实例
                self.calendar_app = CalendarApp(self, start_date, end_date)
                self.calendar_app.show()
            else:
                QMessageBox.warning(self, '输入错误', '请输入正确的日期格式：YYYY.MM.DD。')
        except Exception as e:
            print(f"get_start_date 发生错误: {e}")

    def validate_date(self, start_date_str, end_date_str):  # 验证日期格式是否正确
        try:
            datetime.strptime(start_date_str, "%Y.%m.%d")
            datetime.strptime(end_date_str, "%Y.%m.%d")
            return True
        except ValueError:
            return False

    def is_reasonable_date(self, start_date_obj, end_date_obj):  # 检验两个日期对象是否合理
        try:
            if start_date_obj.year < 2000 or start_date_obj.year > 3000:
                return False
            if end_date_obj.year < 2000 or end_date_obj.year > 3000:
                return False
            return True
        except ValueError:
            return False

    def copy_rename_list_sheets(self, start_date, end_date, file_path):
        print(f"{file_path}")
        try:
            if file_path == "":
                # 打开文件选择对话框
                root = Tk()
                root.withdraw()
                while True:
                    file_path = filedialog.askopenfilename(title="选择模板文件", filetypes=[("Excel files", "*.xlsx")])
                    
            if "模板" in os.path.basename(file_path):
                if "森赫" in os.path.basename(file_path):
                    self.company = "森赫"
                elif "联合" in os.path.basename(file_path):
                    self.company = "联合"
                else:
                    file_path = ""
                    print("选择的文件并非模板文件，请重新选择。")
            else:
                file_path = ""
                print("选择的文件并非模板文件，请重新选择。")

            if file_path != "":
                output_desktop_directory = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop') # 获取当前用户桌面路径
                company_name = "森赫" if self.company == "森赫" else "联合"
                new_filename = f"{company_name}-研发人员工时记录{start_date.strftime('%Y.%m.%d')}-{end_date.strftime('%Y.%m.%d')}.xlsx"
                self.desktop_full_path = os.path.join(output_desktop_directory, new_filename)
                self.start_date_str = start_date.strftime('%Y.%m.%d')
                self.copy_and_rename_template(os.path.dirname(file_path), os.path.basename(file_path), self.desktop_full_path)
            else:
                self.copyButton.setEnabled(False)
        except Exception as e:
            print(f"copy_rename_list_sheets 发生错误: {e}")

    def copy_and_rename_template(self, template_directory, template_filename, new_filename):
        try:
            # 完成复制并重命名逻辑
            shutil.copyfile(os.path.join(template_directory, template_filename), new_filename)
        except Exception as e:
            print(f"复制并重命名模板时发生错误: {e}")
###############################################以下需要调整，否则如果周期天数大于32，则会遗漏数据。
    def update_excel_with_calendar_data(self, excel_file_path, start_date_str, end_date_str, working_days_dict):
        try:
            self.working_days_dict = working_days_dict
            self.excel_info = {
                "excel_file_path": excel_file_path,
                "start_date_str": start_date_str,
                "working_days_dict": working_days_dict
            }
            
            start_date = datetime.strptime(start_date_str, "%Y.%m.%d")
            end_date = datetime.strptime(end_date_str, "%Y.%m.%d")
            xl = openpyxl.load_workbook(excel_file_path)

            fill_none = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type=None)  # 无填充颜色
            fill_blue = PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid")

            # 处理第二个工作表
            second_sheet = xl.worksheets[1]
            for row in range(4, 35):
                for col in range(1, second_sheet.max_column + 10):  # 处理整行
                    cell = second_sheet.cell(row=row, column=col)
                    cell.fill = fill_none
                second_sheet[f'A{row}'].value = None  # 清空A列单元格的内容

            current_date = start_date
            for row in range(4, 35):
                if current_date > end_date:
                    break
                second_sheet[f'A{row}'].value = current_date.strftime('%Y/%m/%d')
                # 判断是否为工作日
                if current_date.strftime('%Y.%m.%d') in working_days_dict:
                    is_working = working_days_dict[current_date.strftime('%Y.%m.%d')]
                    if not is_working:
                        for col in range(1, second_sheet.max_column + 10):  # 处理整行
                            cell = second_sheet.cell(row=row, column=col)
                            cell.fill = fill_blue  # 设置填充颜色为蓝色
                current_date += timedelta(days=1)

            # 设置第二个工作表名称和A1单元格的值
            new_second_sheet_title = f"{end_date.month}月合计"
            print("{end_date.month}")############

            
            second_sheet.title = new_second_sheet_title
            company_full_name = "森赫电梯股份有限公司" if self.company == "森赫" else "浙江联合电梯有限公司"
            second_sheet['A1'].value = f"{company_full_name}{end_date.year}年{end_date.month}月研发人员工时记录"

            # 处理第三个及以后的工作表
            for sheet in xl.worksheets[2:]:
                for row in range(4, 35):
                    for col in range(1, sheet.max_column + 10):  # 处理整行
                        cell = sheet.cell(row=row, column=col)
                        cell.fill = fill_none
                        if col == 1:
                            cell.value = None  # 仅清空A列（第1列）单元格的内容

                current_date = start_date
                for row in range(4, 35):
                    if current_date > end_date:
                        break
                    sheet[f'A{row}'].value = current_date.strftime('%Y/%m/%d')
                    # 判断是否为工作日
                    if current_date.strftime('%Y.%m.%d') in working_days_dict:
                        is_working = working_days_dict[current_date.strftime('%Y.%m.%d')]
                        if not is_working:
                            for col in range(1, sheet.max_column + 10):  # 处理整行
                                cell = sheet.cell(row=row, column=col)
                                cell.fill = fill_blue  # 设置填充颜色为蓝色
                    current_date += timedelta(days=1)

                # 设置每个工作表的A1单元格的值
                sheet['A1'].value = f"{company_full_name}{end_date.year}年{end_date.month}月研发人员工时记录"

                # 更新每个工作表第37行和第39至44行中的公式
                for row in [37, 39, 40, 41, 42, 43, 44]:
                    for col in range(2, sheet.max_column + 10):  # 从B列开始遍历所有列
                        cell = sheet.cell(row=row, column=col)
                        if cell.data_type == 'f':  # 如果单元格是公式
#######################################注意模板的默认公式中的月份是多少
                            if "='1月合计'!" in cell.value:
                                new_formula = cell.value.replace("='1月合计'!", f"='{end_date.month}月合计'!")
                                cell.value = new_formula

            xl.save(excel_file_path)
        except Exception as e:
            print(f"更新Excel文件时发生错误: {e}")

    def list_sheets(self, excel_file_path):
        self.sheetSelector.clear()
        try:
            xl = pd.ExcelFile(excel_file_path)
            for sheet in xl.sheet_names[2:]:
                self.sheetSelector.addItem(sheet)
        except Exception as e:
            print(f"读取工作表失败: {e}")

    def show_a2_content(self, selected_sheet):
        try:
            df = pd.read_excel(self.desktop_full_path, sheet_name=selected_sheet)
            a2_content = df.iloc[0, 0] if not df.empty else '无数据'
            self.a2Label.setText(f'{a2_content}')
        except Exception as e:
            self.a2Label.setText(f'读取失败: {e}')

    def populate_person_selector(self, excel_file_path):
        try:
            xl = pd.ExcelFile(excel_file_path)
            second_sheet = xl.sheet_names[1]  # 第二个工作表
            df = pd.read_excel(excel_file_path, sheet_name=second_sheet)
            persons = []
            for cell in df.iloc[1, 1:]:  # 从A2单元格开始
                if pd.isna(cell):
                    break
                persons.append(str(cell))  # 转换为字符串
            self.personSelector.addItems(persons)
        except Exception as e:
            print(f"读取人员列表失败: {e}")

    def populate_project_inputs(self, projects, days):
        for project, day in zip(projects, days):
            layout = QHBoxLayout()
            label = QLabel(project)  # 创建 QLabel
            input_field = QLineEdit()
            input_field.setText(str(day))
            input_field.setValidator(QIntValidator(0, 100))
            layout.addWidget(label)
            layout.addWidget(input_field)
            self.projectInputLayout.addLayout(layout)

    def tabulate_every1s_work(self, selected_person):
        self.personLabel.setText('')  # 清空当前内容

        try:
            # 从 all_project_data 字典中读取数据
            if selected_person in self.all_project_data:
                project_data = self.all_project_data[selected_person]
                self.project_list = project_data['project_list']
                self.default_days = project_data['default_days']
                self.personLabel.setText('\n'.join(self.project_list))
                print(f'{selected_person.ljust(6 - len(selected_person))} 参与的项目列表: {self.project_list}')
                print(f'{selected_person.ljust(6 - len(selected_person))} 每个项目分配天数: {self.default_days}')
            else:
                print(f"未找到匹配的人员：{selected_person}")

        except Exception as e:
            self.personLabel.setText(f'读取数据失败: {e}')
            print(f'读取数据失败: {e}')
        self.progressBar.setValue(100)
        #print(f"all_project_data: {self.all_project_data}")
        #print(f"working_days_dict: {self.working_days_dict}")

    def tabulate_all_work(self):
        self.personLabel.setText('')  # 清空当前内容

        try:
            print("开始加载工作簿")
            # 加载工作簿一次
            wb = openpyxl.load_workbook(self.desktop_full_path)
            print("工作簿加载完成")
            
            xl = pd.ExcelFile(self.desktop_full_path)
            num_sheets = len(xl.sheet_names[2:])  # 工作表总数
            print(f"工作表总数: {num_sheets}")
            all_project_list = {}  # 用于存储所有人的项目列表

            # 初始化进度条
            progress_step = 84 / num_sheets if num_sheets > 0 else 0
            progress = 12
            self.progressBar.setValue(int(progress))

            # 遍历每个工作表并直接读取数据
            for sheet_index, sheet in enumerate(xl.sheet_names[2:], start=1):  # 从第三个工作表开始统计
                try:
                    print(f"读取工作表: {sheet}")
                    df = pd.read_excel(self.desktop_full_path, sheet_name=sheet)
                    for col_num in range(1, df.shape[1]):
                        cell_value = df.iloc[1, col_num]
                        if pd.isna(cell_value):
                            break
                        #print(f"列号: {col_num}, 值: {cell_value}")
                        if not self.is_column_hidden(wb, sheet, col_num):  # 传递工作簿
                            selected_person = cell_value  # 假设每列对应的人员名称是唯一的
                            
                            print(f"列号：{str(col_num).ljust(3)}，值：{selected_person}")
                            if selected_person not in all_project_list:
                                all_project_list[selected_person] = []
                            all_project_list[selected_person].append(sheet)
                except Exception as e:
                    print(f"处理工作表 {sheet} 时出错: {e}")
                    continue
                
                # 更新进度条
                progress += progress_step
                self.progressBar.setValue(int(progress))

            print("项目列表读取完成")
            self.progressBar.setValue(96)
                
            # 获取所有人员的项目列表
            second_sheet = xl.sheet_names[1]  # 第二个工作表
            try:
                print(f"读取第二个工作表: {second_sheet}")
                df_second = pd.read_excel(self.desktop_full_path, sheet_name=second_sheet)
                projects = []
            except Exception as e:
                print(f"读取第二个工作表 {second_sheet} 时出错: {e}")
                return

            # 获取人名列表，即 Excel 中的 B3, C3, D3...
            try:
                names = df_second.iloc[1, 1:].tolist()  # 第三行索引为1（从0开始计数，并排除掉作为标题行不参与索引计数的第一行）
                print(f"人名列表: {names}")
            except Exception as e:
                print(f"获取人名列表时出错: {e}")
                return

            # 为每个人计算初始分配天数并存储
            for name in names:
                try:
                    selected_index = None
                    for index, person_name in enumerate(names):
                        if person_name == name:
                            selected_index = index + 1
                            break

                    if selected_index is not None:
                        #print(f"处理人: {name}")
                        # 提取对应列的项目数据，从第四行（索引为3）开始
                        projects = df_second.iloc[3:, selected_index].dropna().tolist()
                        projects = [str(item) for item in projects]  # 确保项目数据是字符串

                        # 计算工作日总数
                        work_days = self.get_work_days()

                        # 获取初始分配天数
                        project_list = all_project_list.get(name, [])
                        num_projects = len(project_list)
                        if num_projects > 0:
                            default_days = self.distribute_days_evenly(work_days, num_projects)
                            random.shuffle(default_days)
                        else:
                            default_days = []

                        self.all_project_data[name] = {
                            'project_list': project_list,
                            'default_days': default_days
                        }
                    else:
                        print(f"未找到匹配的列：{name}")
                except Exception as e:
                    print(f"处理人 {name} 时出错: {e}")
        except Exception as e:
            self.personLabel.setText(f'读取数据失败: {e}')
            print(f'读取数据失败: {e}')
        finally:
            self.progressBar.setValue(99)
        self.allocateButton.setEnabled(True)
        self.arrangeButton.setEnabled(True)

    def clear_project_inputs(self):
        print("开始清空项目输入框")
        while self.projectInputLayout.count():
            print(f"布局中的项数: {self.projectInputLayout.count()}")
            item = self.projectInputLayout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                print(f"删除子项: {widget}")
                widget.deleteLater()
        self.projectInputLayout.update()  # 强制更新布局
        print("项目输入框清空完成")
          
    def distribute_days_evenly(self, total_days, num_projects):
        if num_projects == 0:  # 检查项目数量是否为零。如果为零，则直接返回一个空列表，避免除零错误。
            return []        
        base = total_days // num_projects
        remainder = total_days % num_projects
        distribution = [base] * num_projects
        for i in range(remainder):
            distribution[i] += 1
        return distribution

    def apply_project_days(self):
        selected_person = self.personSelector.currentText()
        total_days = sum(1 for date, checkbox in self.calendar_app.working_days if checkbox.isChecked())
        entered_days = []

        for i in range(self.projectInputLayout.count()):
            item = self.projectInputLayout.itemAt(i)
            if item.layout():
                input_field = item.layout().itemAt(1).widget()
                if input_field and isinstance(input_field, QLineEdit):
                    day = input_field.text()
                    if day.isdigit():
                        entered_days.append(int(day))
                    else:
                        QMessageBox.warning(self, '输入错误', f'项目天数必须是正整数：项目 {i + 1}')
                        return

        if sum(entered_days) != total_days:
            QMessageBox.warning(self, '输入错误', f'输入的天数总和必须等于工作日的总天数：{total_days} 天')
            return

        # 将分配的天数应用到选定的人员和项目中
        self.person_project_days[selected_person] = entered_days
        self.clear_project_inputs()
        self.open_project_input_window()


    def is_column_hidden(self, wb, sheet_name, col_num):  # 列宽和隐藏与否检测不出来，只能用单元格背景色来判断
        sheet = wb[sheet_name]
        column_letter = openpyxl.utils.get_column_letter(col_num + 1)
        cell = sheet[f'{column_letter}3']
        fill = cell.fill
        cell_color = fill.start_color.rgb
        green_rgb = "FF92D050"  # RGB(146, 208, 80)的十六进制
        yellow_rgb = "FFFFFF00"  # RGB(255, 255, 80)的十六进制
        
        return (cell_color != green_rgb) and (cell_color != yellow_rgb)

    def open_project_input_window(self):
        work_days = self.get_work_days()
        print(f'Opening ProjectInputWindow with projects: {self.project_list} and days: {self.default_days}')
        self.project_input_window = ProjectInputWindow(self.project_list, self.default_days, work_days)
        self.project_input_window.days_updated.connect(self.update_days)  # 连接信号到槽
        self.project_input_window.show()

    @pyqtSlot(list)
    def update_days(self, days):
        print(f'Updated days: {days}')
        print(f'Previous default_days: {self.default_days}')
        
        # 从 personSelector 获取选中的人名
        selected_person = self.personSelector.currentText()
        print(f'Selected person: {selected_person}')
        
        # 检查 selected_person 是否在 all_project_data 中
        if selected_person not in self.all_project_data:
            print(f'Error: selected_person "{selected_person.ljust(6 - len(selected_person))}" not found in all_project_data')
            return
        
        # 更新 default_days
        self.default_days = days
        print(f'Updated default_days: {self.default_days}')
        
        # 获取并打印当前 selected_person 的默认天数
        previous_days = self.all_project_data[selected_person]['default_days']
        print(f"Previous self.all_project_data['{selected_person.ljust(6 - len(selected_person))}']['default_days']: {previous_days}")
        
        # 更新 all_project_data 中的 selected_person 的默认天数
        self.all_project_data[selected_person]['default_days'] = days
        print(f"Updated self.all_project_data['{selected_person.ljust(6 - len(selected_person))}']['default_days']: from {previous_days} to {days}")
            
    def assign_work_schedule(self):
        self.progressBar2.setValue(0)
        self.allocateButton.setEnabled(False)
        self.arrangeButton.setEnabled(False)
        try:
            if not self.excel_info:
                print("Excel信息未定义")
                return
            excel_file_path = self.excel_info["excel_file_path"]
            working_days_dict = self.excel_info["working_days_dict"]

            # 检查路径和文件合法性
            if not excel_file_path or not os.path.isfile(excel_file_path):
                print(f"无效的文件路径: {excel_file_path}")
                return
            print(f"准备加载工作簿: {excel_file_path}")

            # 加载工作簿
            xl = openpyxl.load_workbook(excel_file_path)
            print(f"工作簿加载成功")
            sheet = xl.worksheets[2]

            # 初始化
            work_schedule = {}
            project_days_left = {}
            self.progressBar2.setValue(1)
            current_progress = 1
            # 森赫需要去掉这几个人：惠志全、何昕鑫、沈凯  、盛嘉媛、钟明明
            num_persons = len(self.all_project_data) - 5 if self.company == "森赫" else len(self.all_project_data)
            num_projects = len(xl.sheetnames[2:])
            num_workdays = sum(1 for date, is_working in working_days_dict.items() if is_working)
            progress_step = 5 / num_projects
            
            for project in xl.sheetnames[2:]:
                project_days_left[project] = {}

                for person, data in self.all_project_data.items():
                    if project in data['project_list']:
                        assigned_days = self.person_project_days.get(person, data['default_days'])
                        project_days_left[project][person] = assigned_days[data['project_list'].index(project)]
                current_progress += progress_step
                self.progressBar2.setValue(int(current_progress))
            self.progressBar2.setValue(6)
            
            # 第一轮: 每个工作日的每个项目分配一个人
            working_days = sorted([date for date, is_working in working_days_dict.items() if is_working])
            current_progress = 6
            progress_step = 18 / num_projects / num_workdays
            for project in xl.sheetnames[2:]:
                rd_workdays = working_days.copy()
                random.shuffle(rd_workdays)  # 重新打乱顺序
                print(f"项目 {project} 打乱后的工作日顺序: {rd_workdays}")

                for date in rd_workdays:
                    if not project_days_left[project]:
                        print(f"项目 {project} 在 {date} 没有剩余天数需要分配，跳过")
                        continue

                    # 选择优先级最高的人，并检查当天是否有安排工作
                    selected_person = max(
                        project_days_left[project],
                        key=lambda person: (
                            project_days_left[project][person] - len(self.all_project_data[person]['project_list'])
                        ) if date not in work_schedule.get(person, {}) else -99
                    )

                    # 检查是否已在这一天有安排工作
                    if selected_person in work_schedule and date in work_schedule[selected_person]:
                        # 打印所有参与人员在当天的 key 值
                        for person in project_days_left[project]:
                            key_value = (project_days_left[project][person] - len(self.all_project_data[person]['project_list'])) * (1 if date not in work_schedule.get(person, {}) else 0)
                            print(f"项目 {project} - 日期 {date} - 人员 {person} - key 值 {key_value}")
                        print(f"项目 {project} 在 {date} 已为 {selected_person} 安排了工作，跳过")
                        continue  # 如果已经有安排工作，跳过

                    if selected_person not in work_schedule:
                        work_schedule[selected_person] = {}
                    work_schedule[selected_person][date] = project
                    project_days_left[project][selected_person] -= 1

                    if project_days_left[project][selected_person] == 0:
                        del project_days_left[project][selected_person]

                    #print(f"为 {selected_person} 在 {date} 分配了项目 {project}")
                    print(f"为 {selected_person.ljust(6 - len(selected_person))} 在 {date} 分配了项目 {project}")
                    current_progress += progress_step
                    self.progressBar2.setValue(int(current_progress))
            self.progressBar2.setValue(24)

            # 第二轮: 依次遍历每个人的每一个工作日，随机分配剩余天数
            current_progress = 24
            progress_step = 48 / num_persons

            for person, schedule in work_schedule.items():
                print(f"开始为{person.ljust(6 - len(person))} 分配项目")
                for date in working_days:
                    if date in schedule:
                        continue  # 如果已有分配项目，跳过

                    # 查找这个人负责的尚未达到参与天数要求的项目
                    eligible_projects = [project for project, days in project_days_left.items() if person in days and days[person] > 0]
                    if eligible_projects:
                        # 随机选一个项目
                        selected_project = random.choice(eligible_projects)
                        work_schedule[person][date] = selected_project
                        project_days_left[selected_project][person] -= 1

                        if project_days_left[selected_project][person] == 0:
                            del project_days_left[selected_project][person]

                        print(f"为 {person.ljust(6 - len(person))}  在 {date} 分配了项目 {selected_project}")

                current_progress += progress_step
                self.progressBar2.setValue(int(current_progress))
            self.progressBar2.setValue(72)
            
            # 打印每个人的工作计划
            current_progress = 72
            progress_step = 28 / num_persons
            for person, schedule in work_schedule.items():
                print(f'工作计划 - {person}:')
                for date, project in sorted(schedule.items()):
                    print(f'  {date}: {project}')
                    
                current_progress += progress_step
                self.progressBar2.setValue(int(current_progress))

            self.progressBar2.setValue(99)

            # 将工作计划存储到 self 中
            self.work_schedule = work_schedule

        except Exception as e:
            print(f"读取或更新Excel文件时发生错误: {e}")
        self.writeButton.setEnabled(True)
        self.progressBar2.setValue(100)


    def write_schedule_2_excel(self):
        self.writeButton.setEnabled(False)
        try:
            if not self.excel_info:
                print("Excel信息未定义")
                return

            excel_file_path = self.excel_info["excel_file_path"]

            # 加载工作簿
            xl = openpyxl.load_workbook(excel_file_path)
            print(f"工作簿加载成功")

            self.progressBar3.setValue(0)
            num_persons = len(self.all_project_data)
            progress_step = 99 / num_persons
            current_progress = 0

            # 更新Excel文件
            for person, schedule in self.work_schedule.items():
                print(f"处理 {person} 的工作计划")

                current_progress += progress_step
                self.progressBar3.setValue(int(current_progress))

                for date, project in sorted(schedule.items()):
                    print(f"日期: {date}, 项目: {project}")

                    if project in xl.sheetnames:
                        sheet = xl[project]

                        column_number = None
                        for col in range(2, sheet.max_column + 1):
                            if sheet.cell(row=3, column=col).value == person:
                                column_number = col
                                break

                        if column_number is None:
                            print(f"没有找到 {person} 的列号")
                            continue

                        row_number = None
                        for row in range(4, 35):
                            cell_value = sheet.cell(row=row, column=1).value
                            if isinstance(cell_value, datetime):
                                cell_date = cell_value.strftime('%Y.%m.%d')
                            else:
                                cell_date = str(cell_value).replace('/', '.')
                            if cell_date == date:
                                row_number = row
                                break

                        if row_number is None:
                            print(f"没有找到 {date} 的行号")
                            continue

                        sheet.cell(row=row_number, column=column_number).value = 8

            self.progressBar3.setValue(99)
            
            print("调整前的工作表顺序：")
            for sheetname in xl.sheetnames:
                print(sheetname)
                
            # 调整工作表顺序
            ordered_sheetnames = xl.sheetnames[:2] + sorted(xl.sheetnames[2:])
            xl._sheets = [xl[sheetname] for sheetname in ordered_sheetnames]
            
            print("调整后的工作表顺序：")
            for sheetname in xl.sheetnames:
                print(sheetname)

            xl.save(excel_file_path)
            print(f"已保存更新后的工作簿 {excel_file_path}")
            self.progressBar3.setValue(100)
        except Exception as e:
            print(f"读取或更新Excel文件时发生错误: {e}")

    def get_work_days(self):
        return sum(is_working for date, is_working in self.calendar_app.working_days)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = ExcelApp()
    sys.exit(app.exec_())
