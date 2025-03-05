import sys
import pandas as pd
import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import QApplication, QWidget, QInputDialog, QVBoxLayout, QPushButton, QComboBox, QListView, QListWidget, QFileDialog, QLabel, QMessageBox, QLineEdit, QDialog, QHBoxLayout

class CustomListView(QListView):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("QListView::item { padding: 3px; }")

class CustomComboBox(QComboBox):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setView(CustomListView(self))

class DateDialog(QDialog):
    def __init__(self, initial_date, parent=None):
        super().__init__(parent)
        self.setWindowTitle('选择年月')
        self.layout = QVBoxLayout()
        self.label = QLabel(initial_date)
        self.layout.addWidget(self.label)

        self.buttonLayout = QHBoxLayout()
        self.prevButton = QPushButton('←')
        self.prevButton.clicked.connect(self.prevMonth)
        self.buttonLayout.addWidget(self.prevButton)

        self.nextButton = QPushButton('→')
        self.nextButton.clicked.connect(self.nextMonth)
        self.buttonLayout.addWidget(self.nextButton)

        self.confirmButton = QPushButton('确定')
        self.confirmButton.clicked.connect(self.accept)
        self.buttonLayout.addWidget(self.confirmButton)

        self.layout.addLayout(self.buttonLayout)
        self.setLayout(self.layout)

        self.date = pd.Timestamp(initial_date)

    def prevMonth(self):
        self.date = self.date - pd.DateOffset(months=1)
        self.label.setText(self.date.strftime('%Y-%m'))

    def nextMonth(self):
        self.date = self.date + pd.DateOffset(months=1)
        self.label.setText(self.date.strftime('%Y-%m'))

    def getDate(self):
        return self.date.strftime('%Y-%m')

class RDMaterialSummaryApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.first_import_click = True
        self.excel_date = pd.Timestamp.now().strftime('%Y-%m')
        self.prev_rd_index = 0  # 用户选择的上一个RD编号，用于回溯
        self.prev_sample_index = 0  # 用户选择的上一个样品编号，用于回溯
        self.material_array = []  # 物料数组，格式为[物料信息，可用次数，由此物料制作的样品编号]
        self.sample_array = []  # 样品数组，格式为[样品编号，样品名称，样品数量，用途与处置方式，使用地点，包含物料，领用日期，处置日期]
        
    def initUI(self):
        self.setWindowTitle('研发领料汇总表')
        self.setGeometry(100, 100, 480, 480)

        self.label = QLabel('请选择本月的研发领料汇总表框架：', self)
        self.label.setGeometry(5, 0, 200, 30)
        
        self.button = QPushButton('选择excel文件', self)
        self.button.setGeometry(205, 4, 270, 22)
        self.button.clicked.connect(self.openAndCopyFile)

        self.label = QLabel('请选择要编辑的研发项目编号：', self)
        self.label.setGeometry(5, 25, 200, 30)
        
        self.comboBox1 = CustomComboBox(self)  # RD下拉框
        self.comboBox1.setGeometry(206, 30, 268, 20)
        self.comboBox1.currentIndexChanged.connect(self.handleRDChange)

        self.label = QLabel('样品编号：', self)
        self.label.setGeometry(5, 50, 80, 30)
        
        self.sampleNumberComboBox = CustomComboBox(self)  # 样品编号下拉框
        self.sampleNumberComboBox.setGeometry(65, 55, 409, 20)
        self.sampleNumberComboBox.currentIndexChanged.connect(self.handleSampleNumberChange)

        self.label = QLabel('样品名称：', self)
        self.label.setGeometry(5, 75, 80, 30)
        
        self.yhpbmkigLineEdit = QLineEdit(self)  # 样品名称文本框
        self.yhpbmkigLineEdit.setGeometry(65, 80, 284, 20)
        self.yhpbmkigLineEdit.textChanged.connect(self.checkFields)

        self.label = QLabel('数量：', self)
        self.label.setGeometry(360, 75, 60, 30)
        
        self.yhpbuullLineEdit = QLineEdit(self)  # 样品数量文本框
        self.yhpbuullLineEdit.setGeometry(400, 80, 74, 20)
        self.yhpbuullLineEdit.textChanged.connect(self.checkFields)

        self.label = QLabel('样品用途与处置方式：', self)
        self.label.setGeometry(5, 100, 120, 30)
        
        self.processComboBox = CustomComboBox(self)  # 处置方式下拉框
        self.processComboBox.setGeometry(205, 105, 269, 20)
        self.processComboBox.addItems(["样品验证-报废", "小批量验证-现场自行处理"])
        self.processComboBox.currentIndexChanged.connect(self.updateLocationComboBox)

        self.label = QLabel('样品使用地点：', self)
        self.label.setGeometry(5, 125, 120, 30)

        self.locationComboBox = CustomComboBox(self)  # 样品使用地点下拉框
        self.locationComboBox.setGeometry(205, 130, 269, 20)

        self.updateLocationComboBox()  # 初始化样品使用地点下拉框选择项

        self.addButton = QPushButton('向列表添加', self)
        self.addButton.setGeometry(5, 154, 110, 22)
        self.addButton.setEnabled(False)
        self.addButton.clicked.connect(self.addToListWidget)

        self.delButton = QPushButton('从列表删除', self)
        self.delButton.setGeometry(125, 154, 110, 22)
        self.delButton.setEnabled(False)
        self.delButton.clicked.connect(self.MinusToListWidget)

        self.plusButton = QPushButton('物料使用次数+1', self)
        self.plusButton.setGeometry(245, 154, 110, 22)
        self.plusButton.setEnabled(False)
        self.plusButton.clicked.connect(self.MaterialPlusOne)

        self.minusButton = QPushButton('物料使用次数-1', self)
        self.minusButton.setGeometry(365, 154, 110, 22)
        self.minusButton.setEnabled(False)
        self.minusButton.clicked.connect(self.MaterialMinusOne)

        self.label = QLabel('物料信息：', self)
        self.label.setGeometry(5, 175, 80, 30)

        self.comboBox2 = CustomComboBox(self)  # 物料下拉框
        self.comboBox2.setGeometry(75, 180, 399, 20)
        
        self.listWidget = QListWidget(self)  # 列表框
        self.listWidget.setGeometry(6, 205, 443, 240)
        self.listWidget.itemSelectionChanged.connect(self.updateButtonStates)

        self.upButton = QPushButton('↑', self)
        self.upButton.setGeometry(453, 204, 22, 120)
        self.upButton.setEnabled(False)
        self.upButton.clicked.connect(self.MaterialUp)

        self.downButton = QPushButton('↓', self)
        self.downButton.setGeometry(453, 326, 22, 120)
        self.downButton.setEnabled(False)
        self.downButton.clicked.connect(self.MaterialDown)

        self.saveButton = QPushButton('记录', self)
        self.saveButton.setGeometry(5, 452, 233, 22)
        self.saveButton.setEnabled(False)
        self.saveButton.clicked.connect(self.saveData)
        
        self.importButton = QPushButton('导入', self)
        self.importButton.setGeometry(242, 452, 230, 22)
        self.importButton.setEnabled(False)
        self.importButton.clicked.connect(self.importData)
        
    def openAndCopyFile(self):
        options = QFileDialog.Options()
        filePath, _ = QFileDialog.getOpenFileName(self, "打开Excel文件", "", "Excel文件 (*.xlsx);;所有文件 (*)", options=options)
        if filePath:
            date_dialog = DateDialog(pd.Timestamp.now().strftime('%Y-%m'), self)
            if date_dialog.exec_() == QDialog.Accepted:
                self.excel_date = date_dialog.getDate()
            self.loadExcelData(filePath)
            self.filePath = filePath
            self.copyFileToDesktop()
            
    def loadExcelData(self, filePath):
        df = pd.read_excel(filePath, sheet_name=0)
        column_name = None
        for col in df.columns:
            if 'RD' in col:
                column_name = col
                break
        
        if column_name:
            unique_values = df[column_name].dropna().unique()
            self.comboBox1.clear()
            self.comboBox1.addItems([""] + list(map(str, unique_values)))
            self.df = df
            self.column_name = column_name
        else:
            self.comboBox1.clear()
            self.comboBox1.addItem("未找到'RD'列")
        
    def copyFileToDesktop(self):
        try:
            if hasattr(self, 'filePath'):
                desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
                self.desktop_file_path = os.path.join(desktop, '15894.xlsx')
                shutil.copy(self.filePath, self.desktop_file_path)
                
                wb = load_workbook(self.desktop_file_path, data_only=True)
                first_sheet = wb.sheetnames[0]
                sheets_to_remove = wb.sheetnames[1:]
                for sheet in sheets_to_remove:
                    wb.remove(wb[sheet])
                
                ws = wb[first_sheet]
                header = {cell.value: col for col, cell in enumerate(ws[1], 1)}
                if '单价' in header and '金额' in header and '数量' in header:
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        if row[header['数量'] - 1].value and row[header['金额'] - 1].value:
                            row[header['单价'] - 1].value = round(row[header['金额'] - 1].value / row[header['数量'] - 1].value, 2)
                        else:
                            row[header['单价'] - 1].value = 0.00
                
                wb.save(self.desktop_file_path)
                
                self.material_array = []
                for idx, row in enumerate(self.df.itertuples(index=False), 2):
                    formatted_row = self.formatRow(row, idx)
                    self.material_array.append([formatted_row, 1, ""])
                
                self.addButton.setEnabled(True)
        except Exception as e:
            self.showError(str(e))
    
    def populateSecondComboBox(self):
        selected_value = self.comboBox1.currentText()
        self.comboBox2.clear()
        if selected_value:
            for item in self.material_array:
                if item[1] - item[2].count('#') > 0:
                    row_num = int(item[0].split('.')[0])
                    if self.df.at[row_num - 2, self.column_name] == selected_value:
                        self.comboBox2.addItem(f"{item[0]}<{item[2].count('#')}/{item[1]}>")
    
    def clearListWidget(self):
        self.listWidget.clear()
    
    def updateSampleNumberComboBox(self):
        selected_value = self.comboBox1.currentText()
        self.sampleNumberComboBox.clear()
        if selected_value:
            base_sample_number = f"{selected_value}/{self.excel_date.replace('-', '')}"
            sample_numbers = [f"{base_sample_number}{str(i).zfill(3)}" for i in range(1, 100)]
            used_sample_numbers = {item[0] for item in self.sample_array}
            available_sample_numbers = [num for num in sample_numbers if num not in used_sample_numbers]
            self.sampleNumberComboBox.addItems(available_sample_numbers)
            self.sampleNumberComboBox.setCurrentIndex(0)
        else:
            self.sampleNumberComboBox.clear()

    def updateLocationComboBox(self):
        selected_process = self.processComboBox.currentText()
        self.locationComboBox.clear()
        if selected_process == "样品验证-报废":
            self.locationComboBox.addItems(["研发中心样品测试区域", "中心实验室", "试验塔"])
            self.locationComboBox.setCurrentIndex(0)
            self.locationComboBox.setEnabled(True)
        elif selected_process == "小批量验证-现场自行处理":
            self.locationComboBox.addItem("手动输入项目名称。。。")
            self.locationComboBox.setCurrentIndex(0)
            self.locationComboBox.setEnabled(False)

    def addToListWidget(self):
        selected_value = self.comboBox2.currentText().split('<')[0]
        sample_number = self.sampleNumberComboBox.currentText()
        if selected_value and not self.isSampleNumberUsed(selected_value, sample_number):
            self.listWidget.addItem(selected_value)
            for item in self.material_array:
                if item[0] == selected_value:
                    item[2] += f"#{sample_number}"
                    print(f"Material Array after addition: {item}")
            self.populateSecondComboBox()
            self.updateButtonStates()
            self.checkFields()
    
    def isSampleNumberUsed(self, material, sample_number):
        for item in self.material_array:
            if item[0] == material and f"#{sample_number}" in item[2]:
                return True
        return False

    def handleRDChange(self):
        if self.listWidget.count() > 0:
            response = QMessageBox.warning(self, "警告", "有数据未保存，是否切换RD编号", QMessageBox.Yes | QMessageBox.No)
            if response == QMessageBox.No:
                self.comboBox1.blockSignals(True)
                self.comboBox1.setCurrentIndex(self.prev_rd_index)
                self.comboBox1.blockSignals(False)
                return
            else:
                self.clearListWidget()
                self.clearMaterialArraySampleNumbers()
        
        self.prev_rd_index = self.comboBox1.currentIndex()
        self.populateSecondComboBox()
        self.updateSampleNumberComboBox()
        self.checkFields()

    def handleSampleNumberChange(self):
        if self.listWidget.count() > 0:
            response = QMessageBox.warning(self, "警告", "有数据未保存，是否切换样品编号", QMessageBox.Yes | QMessageBox.No)
            if response == QMessageBox.No:
                self.sampleNumberComboBox.blockSignals(True)
                self.sampleNumberComboBox.setCurrentIndex(self.prev_sample_index)
                self.sampleNumberComboBox.blockSignals(False)
                return
            else:
                self.clearListWidget()
                self.clearMaterialArraySampleNumbers()
        
        self.prev_sample_index = self.sampleNumberComboBox.currentIndex()
        self.checkFields()

    def clearMaterialArraySampleNumbers(self):
        sample_number = self.sampleNumberComboBox.currentText()
        for item in self.material_array:
            item[2] = item[2].replace(f"#{sample_number}", "")

    def MinusToListWidget(self):
        selected_items = self.listWidget.selectedItems()
        if selected_items:
            sample_number = self.sampleNumberComboBox.currentText()
            for item in selected_items:
                item_text = item.text()
                self.listWidget.takeItem(self.listWidget.row(item))
                for material in self.material_array:
                    if material[0] == item_text:
                        material[2] = material[2].replace(f"#{sample_number}", "")
                        print(f"Material Array after deletion: {material}")
            self.populateSecondComboBox()
            self.updateButtonStates()
            self.checkFields()

    def MaterialPlusOne(self):
        selected_items = self.listWidget.selectedItems()
        if selected_items:
            for item in selected_items:
                item_text = item.text()
                for material in self.material_array:
                    if material[0] == item_text and material[1] < 5:
                        material[1] += 1
                        print(f"Material Array after +1: {material}")
            self.populateSecondComboBox()
            self.updateButtonStates()
            self.checkFields()

    def MaterialMinusOne(self):
        selected_items = self.listWidget.selectedItems()
        if selected_items:
            for item in selected_items:
                item_text = item.text()
                for material in self.material_array:
                    if material[0] == item_text and material[1] > 1:
                        material[1] -= 1
                        print(f"Material Array after -1: {material}")
            self.populateSecondComboBox()
            self.updateButtonStates()
            self.checkFields()

    def MaterialUp(self):
        selected_items = self.listWidget.selectedItems()
        if selected_items:
            for item in selected_items:
                row = self.listWidget.row(item)
                if row > 0:
                    self.listWidget.takeItem(row)
                    self.listWidget.insertItem(row - 1, item)
                    self.listWidget.setCurrentItem(item)
        self.checkFields()

    def MaterialDown(self):
        selected_items = self.listWidget.selectedItems()
        if selected_items:
            for item in selected_items:
                row = self.listWidget.row(item)
                if row < self.listWidget.count() - 1:
                    self.listWidget.takeItem(row)
                    self.listWidget.insertItem(row + 1, item)
                    self.listWidget.setCurrentItem(item)
        self.checkFields()

    def formatRow(self, row, idx):
        序号 = f"{idx}"
        品名 = getattr(row, '品名', '')[:36]
        规格 = getattr(row, '规格', '')[:24]
        数量 = getattr(row, '数量', '')
        单位 = getattr(row, '单位', '')
        return f"{序号}.{品名}({规格})_{数量} {单位}"

    def saveData(self):
        try:
            # 检查是否需要手动输入项目名称
            if self.locationComboBox.currentText() == "手动输入项目名称。。。":
                project_name, ok = QInputDialog.getText(self, "输入项目名称", "请输入项目名称：")
                if ok and project_name:
                    self.locationComboBox.setItemText(0, project_name)
                else:
                    return  # 用户取消输入，退出函数
            
            # 获取控件数据
            sample_number = self.sampleNumberComboBox.currentText()
            sample_name = self.yhpbmkigLineEdit.text()
            sample_quantity = self.yhpbuullLineEdit.text()
            process_mode = self.processComboBox.currentText()
            usage_location = self.locationComboBox.currentText()
            material_list = [self.listWidget.item(i).text().split('.')[0] for i in range(self.listWidget.count())]  # 只记录物料的序号

            # 确定领用日期
            rd_project = sample_number.split('/')[0]
            sample_suffix = int(sample_number.split('/')[-1])
            if sample_suffix == 1:
                # 第一个样品
                year_month = self.excel_date
                day = random.randint(1, 3)
                sample_date = pd.Timestamp(f"{year_month}-{day:02d}")
            else:
                # 非第一个样品
                prev_sample_number = f"{rd_project}/{self.excel_date.replace('-', '')}{str(sample_suffix - 1).zfill(3)}"
                prev_sample_date = next((item[6] for item in self.sample_array if item[0] == prev_sample_number), None)
                if prev_sample_date:
                    sample_date = pd.Timestamp(prev_sample_date) + pd.DateOffset(days=random.randint(0, 3))
                else:
                    sample_date = pd.Timestamp.now()
            
            # 确定处置日期
            if usage_location == "中心实验室":
                disposal_date = sample_date + pd.DateOffset(months=3)
            elif usage_location == "研发中心样品测试区域":
                disposal_date = sample_date + pd.DateOffset(months=2)
            elif usage_location == "试验塔":
                disposal_date = sample_date + pd.DateOffset(months=6)
            else:
                disposal_date = None

            # 生成self.sample_array元素
            sample_array_element = [
                sample_number,
                sample_name,
                sample_quantity,
                process_mode,
                usage_location,
                material_list,
                sample_date.strftime('%Y-%m-%d'),  # 领用日期
                disposal_date.strftime('%Y-%m-%d') if disposal_date else ''  # 处置日期
            ]
            self.sample_array.append(sample_array_element)

            # Print生成的样品信息
            print("生成的样品信息:", sample_array_element)

            # 更新self.material_array中的样品编号
            for item in material_list:
                for material in self.material_array:
                    if material[0].split('.')[0] == item:
                        material[2] += f"#{sample_number}"

            # 清空控件
            self.yhpbmkigLineEdit.clear()
            self.yhpbuullLineEdit.clear()
            self.listWidget.clear()
            self.sampleNumberComboBox.setCurrentIndex(0)

            # 更新按钮状态
            self.checkFields()
        except Exception as e:
            self.showError(str(e))
    
    def importData(self):
        try:
            if hasattr(self, 'desktop_file_path'):
                df = pd.read_excel(self.desktop_file_path, sheet_name=0)
                sheet_name = df.columns[0] + "用途及最终去向（研发）"
                
                wb = load_workbook(self.desktop_file_path)
                if self.first_import_click:
                    ws = wb.create_sheet(title=sheet_name)
                    header = df.columns.tolist()
                    for col_num, header_value in enumerate(header, 1):
                        col_letter = get_column_letter(col_num)
                        ws[f"{col_letter}1"].value = header_value
                    self.first_import_click = False
                else:
                    ws = wb[sheet_name]
                    
                for row in range(self.listWidget.count()):
                    item = self.listWidget.item(row).text()
                    if item:
                        row_num = int(item.split('.')[0])
                        row_data = df.iloc[row_num - 2].values.tolist()
                        self.copy_row(ws, row_data)
                    
                wb.save(self.desktop_file_path)
        except Exception as e:
            self.showError(str(e))
    
    def copy_row(self, ws, row_data):
        row_num = ws.max_row + 1
        for idx, value in enumerate(row_data):
            col_letter = get_column_letter(idx + 1)
            cell = ws[f"{col_letter}{row_num}"]
            cell.value = value
    
    def showError(self, message):
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Critical)
        msgBox.setText("发生错误")
        msgBox.setInformativeText(message)
        msgBox.setWindowTitle("错误")
        msgBox.exec_()

    def checkFields(self):
        if (self.listWidget.count() > 0 and self.sampleNumberComboBox.currentText() and 
            self.yhpbmkigLineEdit.text() and self.yhpbuullLineEdit.text()):
            self.importButton.setEnabled(True)
            self.saveButton.setEnabled(True)
        else:
            self.importButton.setEnabled(False)
            self.saveButton.setEnabled(False)

    def updateButtonStates(self):
        has_selection = len(self.listWidget.selectedItems()) > 0
        self.delButton.setEnabled(has_selection)
        self.plusButton.setEnabled(has_selection and self.getSelectedMaterialCount() < 5)
        self.minusButton.setEnabled(has_selection and self.getSelectedMaterialCount() > 1)
        self.upButton.setEnabled(has_selection and self.listWidget.currentRow() > 0)
        self.downButton.setEnabled(has_selection and self.listWidget.currentRow() < self.listWidget.count() - 1)

    def getSelectedMaterialCount(self):
        selected_items = self.listWidget.selectedItems()
        if selected_items:
            item_text = selected_items[0].text()
            for material in self.material_array:
                if material[0] == item_text:
                    return material[1]
        return 0

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = RDMaterialSummaryApp()
    ex.show()
    sys.exit(app.exec_())
